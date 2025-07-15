import os
import time
import logging
from datetime import datetime
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import traceback
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import xlrd
import xlwt
from xlutils.copy import copy


URL_LOGIN = 'https://agenciapublicadeempleo.sena.edu.co/spe-web/spe/login'
URL_FORMULARIO = 'https://ape.sena.edu.co/spe-web/spe/funcionario/empresa'
    
# 1. Selector de la tabla principal
TABLA_SELECTOR = "table#bus-table"

# 2. Selectores para encabezados
ENCABEZADOS_SELECTOR = "thead tr th"

# 3. Selectores para filas de datos
FILAS_DATOS_SELECTOR = "tbody tr"

# 4. Selectores para celdas
CELDAS_SELECTOR = "td"

# 5. Selector para botón "Siguiente"
BOTON_SIGUIENTE_SELECTOR = "a#bus-table_next"

# 6. Selector para botones de acción en cada fila
BOTON_ACCION_SELECTOR = "td button"

# 7. Selectores para información específica de tu tabla
SELECTORES_COLUMNAS = {
    "tipo_id": "td:nth-child(1)",
    "identificacion": "td:nth-child(2)", 
    "nombre_empresa": "td:nth-child(3)",
    "actividad_economica": "td:nth-child(4)",
    "fecha_inscripcion": "td:nth-child(5)",
    "estado": "td:nth-child(6)",
    "accion": "td:nth-child(7)"
}



# Configurar logging
log_filename = f"automatizacion_aprendices_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"
logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Cargar variables de entorno
load_dotenv()


# --- Credenciales de login desde variables de entorno ---
USUARIO_LOGIN = os.getenv('USUARIO_LOGIN')
CONTRASENA_LOGIN = os.getenv('CONTRASENA_LOGIN')
if not USUARIO_LOGIN or not CONTRASENA_LOGIN:
    error_msg = "Error: Las credenciales de login no están configuradas en el archivo .env"
    logging.error(error_msg)
    print(error_msg)
    exit()

class ExtractorDatosEmpresa:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.datos_extraidos = []
        
    def configurar_driver(self):
        """
        Configura el driver de Chrome con opciones optimizadas
        """
        try:
            chrome_options = Options()
            # Opciones para mejor rendimiento
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_argument('--disable-images')  # No cargar imágenes para mayor velocidad
            chrome_options.add_argument('--disable-javascript')  # Solo si la página funciona sin JS
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # Configurar descarga automática
            prefs = {
                "profile.default_content_setting_values.notifications": 2,
                "profile.default_content_settings.popups": 0,
            }
            chrome_options.add_experimental_option("prefs", prefs)
            
            # Usar ChromeDriverManager para manejo automático del driver
            service = ChromeService(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            # Configurar tiempo de espera
            self.wait = WebDriverWait(self.driver, 10)
            self.driver.implicitly_wait(5)
            
            logging.info("Driver configurado exitosamente")
            return True
            
        except Exception as e:
            logging.error(f"Error configurando driver: {str(e)}")
            return False
    
    def realizar_login(self):
        """
        Realiza el login en la plataforma
        """
        """Realiza el proceso de login en la aplicación"""
        try:

            self.driver.get(URL_LOGIN)
            logging.info("Abriendo página de login...")
            
            # Esperar a que el radio button esté disponible
            radio_persona = self.wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//input[@type='radio' and @name='tipousuario' and @value='0']")
            ))
            radio_persona.click()
            
            # Esperar y completar campos de login
            campo_usuario = self.wait.until(EC.presence_of_element_located((By.ID, 'username')))
            campo_usuario.clear()
            campo_usuario.send_keys(USUARIO_LOGIN)
            
            campo_contrasena = self.driver.find_element(By.ID, 'password')
            campo_contrasena.clear()
            campo_contrasena.send_keys(CONTRASENA_LOGIN)
            
            boton_entrar = self.driver.find_element(By.ID, 'entrar')
            boton_entrar.click()
            
            # Esperar a que se complete el login
            self.wait.until(EC.url_changes(URL_LOGIN))
            logging.info("Login exitoso")
            return True
            
        except Exception as e:
            logging.error(f"Error durante el login: {str(e)}")
            print(f"Error durante el login: {str(e)}")
            return False
        


    # FUNCIÓN ADAPTADA PARA TU TABLA ESPECÍFICA
    def extraer_datos_tabla_especifica(self):
        """
        Función específica para extraer datos de tu tabla
        """
        try:
            # Esperar a que la tabla esté presente
            tabla = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, TABLA_SELECTOR))
            )
            
            # Extraer encabezados específicos
            encabezados = []
            headers = tabla.find_elements(By.CSS_SELECTOR, ENCABEZADOS_SELECTOR)
            
            for header in headers:
                texto = header.text.strip()
                if texto:  # Solo agregar si no está vacío
                    encabezados.append(texto)
            
            # Extraer datos de las filas
            filas = tabla.find_elements(By.CSS_SELECTOR, FILAS_DATOS_SELECTOR)
            datos_pagina = []
            
            for fila in filas:
                try:
                    # Extraer datos específicos de cada columna
                    fila_datos = {}
                    
                    # Tipo de ID
                    try:
                        nit = fila.find_element(By.CSS_SELECTOR, "td:nth-child(1)").text.strip()
                        fila_datos["Tipo Id."] = nit
                    except:
                        fila_datos["Tipo Id."] = "N/A"
                    
                    # Identificación
                    try:
                        empresa = fila.find_element(By.CSS_SELECTOR, "td:nth-child(2)").text.strip()
                        fila_datos["Identificación"] = empresa
                    except:
                        fila_datos["Identificación"] = "N/A"
                    
                    # Nombre Empresa
                    try:
                        construccion = fila.find_element(By.CSS_SELECTOR, "td:nth-child(3)").text.strip()
                        fila_datos["Nombre Empresa"] = construccion
                    except:
                        fila_datos["Nombre Empresa"] = "N/A"
                    
                    # Actividad Economica
                    try:
                        fecha = fila.find_element(By.CSS_SELECTOR, "td:nth-child(4)").text.strip()
                        fila_datos["Actividad Económica"] = fecha
                    except:
                        fila_datos["Actividad Económica"] = "N/A"
                    
                    # Fecha de inscripción
                    try:
                        estado = fila.find_element(By.CSS_SELECTOR, "td:nth-child(5)").text.strip()
                        fila_datos["Fecha de Inscripción"] = estado
                    except:
                        fila_datos["Fecha de Inscripción"] = "N/A"
                    
                    #Estado
                    try:
                        estado = fila.find_element(By.CSS_SELECTOR, "td:nth-child(6)").text.strip()
                        fila_datos["Estado"] = estado
                    except:
                        fila_datos["Estado"] = "N/A"

                    # Acción
                    try:
                        accion = fila.find_element(By.CSS_SELECTOR, "td:nth-child(7)").text.strip()
                        fila_datos["Acción"] = accion
                    except:
                        fila_datos["Acción"] = "N/A"
                    
                    datos_pagina.append(fila_datos)
                    
                except Exception as e:
                    logging.warning(f"Error procesando fila: {str(e)}")
                    continue
            
            return datos_pagina
            
        except Exception as e:
            logging.error(f"Error extrayendo datos de tabla: {str(e)}")
        return []
 
    def navegar_paginas(self):
        """
        Navega por todas las páginas de la tabla
        """
        try:
            pagina_actual = 1
            
            while True:
                logging.info(f"Procesando página {pagina_actual}")
                
                # Extraer datos de la página actual
                datos_pagina = self.extraer_datos_tabla_especifica()
                self.datos_extraidos.extend(datos_pagina)
                
                # Buscar botón "Siguiente"
                try:
                    boton_siguiente = self.driver.find_element(By.ID, "bus-table_next")
                    
                    # Verificar si el botón está habilitado
                    if "disabled" in boton_siguiente.get_attribute("class"):
                        logging.info("Llegamos a la última página")
                        break
                    
                    # Hacer click en siguiente
                    self.driver.execute_script("arguments[0].click();", boton_siguiente)
                    time.sleep(2)  # Esperar a que cargue la nueva página
                    
                    pagina_actual += 1
                    
                except Exception as e:
                    logging.info("No se encontró botón siguiente o error navegando")
                    break
            
            logging.info(f"Total de registros extraídos: {len(self.datos_extraidos)}")
            return True
            
        except Exception as e:
            logging.error(f"Error navegando páginas: {str(e)}")
            return False
    
    def crear_archivo_excel(self, nombre_archivo=None):
        """
        Crea un archivo Excel con los datos extraídos usando pandas y openpyxl
        """
        try:
            if not self.datos_extraidos:
                logging.error("No hay datos para exportar")
                return False
            
            # Crear nombre del archivo si no se proporciona
            if not nombre_archivo:
                timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                nombre_archivo = f"datos_empresas_{timestamp}.xlsx"
            
            # Crear DataFrame de pandas
            df = pd.DataFrame(self.datos_extraidos)
            
            # Crear archivo Excel con pandas
            with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Datos_Empresas', index=False)
            
            # Aplicar formato con openpyxl
            self.aplicar_formato_excel(nombre_archivo)
            
            logging.info(f"Archivo Excel creado exitosamente: {nombre_archivo}")
            print(f"✅ Archivo Excel creado: {nombre_archivo}")
            print(f"📊 Total de registros: {len(self.datos_extraidos)}")
            
            return True
            
        except Exception as e:
            logging.error(f"Error creando archivo Excel: {str(e)}")
            return False
    
    def aplicar_formato_excel(self, nombre_archivo):
        """
        Aplica formato profesional al archivo Excel
        """
        try:
            # Cargar el archivo
            wb = load_workbook(nombre_archivo)
            ws = wb.active
            
            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar formato a encabezados
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar bordes a toda la tabla
            from openpyxl.styles import Border, Side
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Guardar cambios
            wb.save(nombre_archivo)
            
        except Exception as e:
            logging.error(f"Error aplicando formato: {str(e)}")
            
    def navegar_a_empresas(self):
        """
        Navega específicamente a la página de empresas
        """
        try:
            print("Navagando a Empresas")
            # Variante 2: Buscar por clase y texto
            dropdown_empresa = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[@class='dropdown-toggle' and contains(text(), 'Empresa')]"))
            )  
            print("Encontrado")
            dropdown_empresa.click()
            
            time.sleep(2)
            # Esperar a que el submenú esté visible
            opcion_registro = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//ul[@class='dropdown-menu']//a[contains(text(), 'Registro y/o actualización de empresa')]"))
            )
            print("opcion encontrada")
            opcion_registro.click()
            
            time.sleep(2)
            
            # 5. Verificar que llegamos a la página correcta
            current_url = self.driver.current_url
            print(f"URL actual: {current_url}")
            
            if "funcionario/empresa" in current_url:
                print("✅ Navegación exitosa via menú")
                return True
            else:
                print("❌ No se pudo navegar via menú")
                return False
        except Exception as e:
            logging.error(f"Error navegando via menú: {str(e)}")
            print(f"❌ Error en navegación via menú: {str(e)}")
            return False
        
    def hacer_consulta_avanzada(self, municipio):
        
        print("haciendo busqueda avanzada")
        boton_avanzada = self.wait.until(
            EC.element_to_be_clickable((By.ID, "btnBusqAvanzada"))
        )
        boton_avanzada.click()
        time.sleep(3)
        
        # Esperar a que el modal esté visible
        self.wait.until(
            EC.visibility_of_element_located((By.ID, "modalBusqAvanzada"))
        )
        
        # Llenar el campo de pais 
        input_pais = self.wait.until(EC.presence_of_element_located((By.ID,"bus-pais")))
        selctor_pais = Select(input_pais)
        selctor_pais.select_by_visible_text('Colombia')
        print("Se selecciono Colombia en el input")
        
        # Llenar el campo de departamento
        input_departamento = self.wait.until(EC.presence_of_element_located((By.ID, "bus-departamento")))
        selector_departamento = Select(input_departamento)
        selector_departamento.select_by_visible_text('Cundinamarca')
        print("Se selecciono Cundinamarca en departamenta")
        
        # Llenar campo con los municipios que lleguen como paramentro
        input_municipio = self.wait.until(EC.presence_of_element_located((By.ID, "bus-municipio")))
        Select(input_municipio).select_by_visible_text(municipio)
        print(f"Se selecciono el municipio {municipio}")
        time.sleep(2)
        
        boton_buscar = self.wait.until(EC.element_to_be_clickable((By.ID, "btnBuscarAv")))
        boton_buscar.click()
        time.sleep(2)
        
        
        return True
    
    def ejecutar_proceso_completo(self, municipios):
        """
        Ejecuta todo el proceso de extracción
        """
        try:
            print("🚀 Iniciando proceso de extracción de datos...")
            
            # Configurar driver
            if not self.configurar_driver():
                return False
            
            # Realizar login
            if not self.realizar_login():
                return False
            
            time.sleep(5)
            
            # Realizar navegacion
            if not self.navegar_a_empresas():
                return False
            
            for municipio in municipios:
                print(f"Municipio a procesar {municipio}")
                if not self.hacer_consulta_avanzada(municipio):
                    logging.warning(f"No se pudo realizar la consulta para el municipio --> {municipio}")
                    continue
                
                try:
                    self.wait.until(EC.presence_of_element_located((By.ID, "bus-table")))
                    print("📄 Tabla encontrada, extrayendo datos...")
                    
                    # Aquí agregas la extracción de todas las filas
                    self.extraer_datos_tabla_especifica()
                except Exception as e:
                    print("Error al buscar el municipio ", e)
            try:    
                # Esperar a que aparezca algún elemento característico de esa página
                self.wait.until(
                    EC.presence_of_element_located((By.ID, "bus-table"))  # O el elemento que confirme que estás en la página correcta
                )
                logging.info("Navegación a página de empresas exitosa")
            except:
                logging.warning("No se pudo confirmar la navegación a la página de empresas")
            
            # Crear archivo Excel
            if not self.crear_archivo_excel():
                return False
            
            print("✅ Proceso completado exitosamente!")
            return True
            
        except Exception as e:
            logging.error(f"Error en proceso completo: {str(e)}")
            print(f"❌ Error en el proceso: {str(e)}")
            return False
        
        finally:
            # Cerrar driver
            if self.driver:
                self.driver.quit()
                logging.info("Driver cerrado")

def main():
    """
    Función principal
    """
    try:
        municipios = [
            "Mosquera",
            "Madrid",
            "Bojaca",
            "Zipacon",
            "Facatativá",
            "El Rosal",
            "Subachoqe",
            "Tabio",
            "Cota",
            "Funza",
            "Tenjo",
            "Guasca",
            "Gacheta",
            "Gama",
            "Ubala",
            "Gachala",
            "Junin"
        ]
        
        
        extractor = ExtractorDatosEmpresa()
        extractor.ejecutar_proceso_completo(municipios)
        
    except KeyboardInterrupt:
        print("\n⚠️  Proceso interrumpido por el usuario")
        logging.info("Proceso interrumpido por el usuario")
    
    except Exception as e:
        print(f"❌ Error inesperado: {str(e)}")
        logging.error(f"Error inesperado: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    main()
