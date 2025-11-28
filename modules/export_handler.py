"""
Módulo de exportación de datos a Excel
"""
import logging
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

from config import settings


class ExportHandler:
    """Maneja la exportación de datos a Excel"""
    
    @staticmethod
    def crear_archivo_excel(datos, nombre_archivo):
        """
        Crea un archivo Excel con los datos proporcionados
        
        Args:
            datos: Lista de diccionarios con los datos
            nombre_archivo: Nombre del archivo a crear
            
        Returns:
            dict: Siempre retorna un diccionario con el resultado
                {"status": "ok"|"nodata"|"error", "path": str, "message": str}
        """
        reports_dir = settings.get_reports_dir()
        ruta_archivo = os.path.join(reports_dir, nombre_archivo)

        try:
            # Validar que hay datos
            if not datos:
                return {
                    "status": "nodata",
                    "path": None,
                    "message": "No hay datos para exportar"
                }
            
            # Crear DataFrame
            df = pd.DataFrame(datos)

            # Crear archivo Excel
            with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Datos_Empresas', index=False)
            
            # Aplicar formato
            ExportHandler._aplicar_formato_excel(ruta_archivo)
            
            return {
                "status": "ok",
                "path": ruta_archivo,
                "message": "Archivo creado exitosamente"
            }
            
        except Exception as e:
            logging.error(f"Error creando archivo Excel: {str(e)}")
            return {
                "status": "error",
                "path": None,
                "message": str(e)
            }
    
    
    @staticmethod
    def _aplicar_formato_excel(nombre_archivo):
        """
        Aplica formato profesional al archivo Excel
        
        Args:
            nombre_archivo: Ruta del archivo Excel
        """
        try:
            # Cargar el archivo
            wb = load_workbook(nombre_archivo)
            ws = wb.active
            
            # Definir estilos
            header_fill = PatternFill(
                start_color=settings.EXCEL_STYLES["header_color"],
                end_color=settings.EXCEL_STYLES["header_color"],
                fill_type="solid"
            )
            header_font = Font(
                color=settings.EXCEL_STYLES["header_font_color"],
                bold=True
            )
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar formato a encabezados
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            
            # Ajustar ancho de columnas
            ExportHandler._ajustar_columnas(ws)
            
            # Aplicar bordes
            ExportHandler._aplicar_bordes(ws)
            
            # Guardar cambios
            wb.save(nombre_archivo)
            
        except Exception as e:
            logging.error(f"Error aplicando formato: {str(e)}")
    
    @staticmethod
    def _ajustar_columnas(worksheet):
        """
        Ajusta el ancho de las columnas según el contenido
        
        Args:
            worksheet: Hoja de trabajo de openpyxl
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(
                max_length + 2,
                settings.EXCEL_STYLES["max_column_width"]
            )
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _aplicar_bordes(worksheet):
        """
        Aplica bordes a todas las celdas de la tabla
        
        Args:
            worksheet: Hoja de trabajo de openpyxl
        """
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border