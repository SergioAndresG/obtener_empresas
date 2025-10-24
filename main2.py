import os
import time
import logging
from datetime import datetime, timedelta
from httpcore import TimeoutException
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import traceback
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from xlutils.copy import copy


URL_LOGIN = 'https://agenciapublicadeempleo.sena.edu.co/spe-web/spe/login'
URL_FORMULARIO = 'https://ape.sena.edu.co/spe-web/spe/funcionario/empresa'
    
# 1. Selector de la tabla principal
TABLA_SELECTOR = "table#bus-table"

# 2. Selectores para encabezados
ENCABEZADOS_SELECTOR = "thead tr th"

# 3. Selectores para filas de datos
FILAS_DATOS_SELECTOR = "tbody tr"

# 4. Selectores para celdas
CELDAS_SELECTOR = "td"

# 5. Selector para botón "Siguiente"
BOTON_SIGUIENTE_SELECTOR = "a#bus-table_next"

# 6. Selector para botones de acción en cada fila
BOTON_ACCION_SELECTOR = "td button"

# 7. Selectores para información específica de tu tabla
SELECTORES_COLUMNAS = {
    "tipo_id": "td:nth-child(1)",
    "identificacion": "td:nth-child(2)", 
    "nombre_empresa": "td:nth-child(3)",
    "actividad_economica": "td:nth-child(4)",
    "fecha_inscripcion": "td:nth-child(5)",
    "estado": "td:nth-child(6)",
    "accion": "td:nth-child(7)"
}



# Configurar logging
log_filename = f"automatizacion_aprendices_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"
logging.basicConfig(
    filename=log_filename,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Cargar variables de entorno
load_dotenv()


# --- Credenciales de login desde variables de entorno ---
USUARIO_LOGIN = os.getenv('USUARIO_LOGIN')
CONTRASENA_LOGIN = os.getenv('CONTRASENA_LOGIN')
if not USUARIO_LOGIN or not CONTRASENA_LOGIN:
    error_msg = "Error: Las credenciales de login no están configuradas en el archivo .env"
    logging.error(error_msg)
    print(error_msg)
    exit()

class ExtractorDatosEmpresa:
    def __init__(self):
        self.driver = None
        self.wait = None
        self.datos_extraidos = []
        
    def configurar_driver(self):
        """
        Configura el driver de Chrome con opciones optimizadas
        """
        try:
            chrome_options = Options()
            # Opciones para mejor rendimiento
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_argument('--disable-images')  # No cargar imágenes para mayor velocidad
            chrome_options.add_argument('--disable-javascript')  # Solo si la página funciona sin JS
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # Configurar descarga automática
            prefs = {
                "profile.default_content_setting_values.notifications": 2,
                "profile.default_content_settings.popups": 0,
            }
            chrome_options.add_experimental_option("prefs", prefs)
            
            # Usar ChromeDriverManager para manejo automático del driver
            service = ChromeService(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            # Configurar tiempo de espera
            self.wait = WebDriverWait(self.driver, 10)
            self.driver.implicitly_wait(5)
            
            logging.info("Driver configurado exitosamente")
            return True
            
        except Exception as e:
            logging.error(f"Error configurando driver: {str(e)}")
            return False
    
    def realizar_login(self):
        """
        Realiza el login en la plataforma
        """
        """Realiza el proceso de login en la aplicación"""
        try:

            self.driver.get(URL_LOGIN)
            logging.info("Abriendo página de login...")
            

            self.wait.until(EC.invisibility_of_element_located((By.ID, "content-load")))


            # Esperar a que el radio button esté disponible
            radio_persona = self.wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//input[@type='radio' and @name='tipousuario' and @value='0']")
            ))
            radio_persona.click()
            
            # Esperar y completar campos de login
            campo_usuario = self.wait.until(EC.presence_of_element_located((By.ID, 'username')))
            campo_usuario.clear()
            campo_usuario.send_keys(USUARIO_LOGIN)
            
            campo_contrasena = self.driver.find_element(By.ID, 'password')
            campo_contrasena.clear()
            campo_contrasena.send_keys(CONTRASENA_LOGIN)
            
            boton_entrar = self.driver.find_element(By.ID, 'entrar')
            boton_entrar.click()
            
            # Esperar a que se complete el login
            self.wait.until(EC.url_changes(URL_LOGIN))
            logging.info("Login exitoso")
            return True
            
        except Exception as e:
            logging.error(f"Error durante el login: {str(e)}")
            print(f"Error durante el login: {str(e)}")
            return False
        


    # FUNCIÓN ADAPTADA PARA TU TABLA ESPECÍFICA
    def extraer_datos_tabla_especifica(self):
        """
        Función específica para extraer datos de tu tabla
        """
        try:
            # Esperar a que la tabla esté presente
            tabla = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, TABLA_SELECTOR))
            )
            
            # Extraer encabezados específicos
            encabezados = []
            headers = tabla.find_elements(By.CSS_SELECTOR, ENCABEZADOS_SELECTOR)
            
            for header in headers:
                texto = header.text.strip()
                if texto:  # Solo agregar si no está vacío
                    encabezados.append(texto)
            
            # Extraer datos de las filas
            filas = tabla.find_elements(By.CSS_SELECTOR, FILAS_DATOS_SELECTOR)
            datos_pagina = []
            
            for fila in filas:
                try:
                    # Extraer datos específicos de cada columna
                    fila_datos = {}
                    
                    # Tipo de ID
                    try:
                        nit = fila.find_element(By.CSS_SELECTOR, "td:nth-child(1)").text.strip()
                        fila_datos["Tipo Id."] = nit
                    except:
                        fila_datos["Tipo Id."] = "N/A"
                    
                    # Identificación
                    try:
                        empresa = fila.find_element(By.CSS_SELECTOR, "td:nth-child(2)").text.strip()
                        fila_datos["Identificación"] = empresa
                    except:
                        fila_datos["Identificación"] = "N/A"
                    
                    # Nombre Empresa
                    try:
                        construccion = fila.find_element(By.CSS_SELECTOR, "td:nth-child(3)").text.strip()
                        fila_datos["Nombre Empresa"] = construccion
                    except:
                        fila_datos["Nombre Empresa"] = "N/A"
                    
                    # Actividad Economica
                    try:
                        fecha = fila.find_element(By.CSS_SELECTOR, "td:nth-child(4)").text.strip()
                        fila_datos["Actividad Económica"] = fecha
                    except:
                        fila_datos["Actividad Económica"] = "N/A"
                    
                    # Fecha de inscripción
                    try:
                        estado = fila.find_element(By.CSS_SELECTOR, "td:nth-child(5)").text.strip()
                        fila_datos["Fecha de Inscripción"] = estado
                    except:
                        fila_datos["Fecha de Inscripción"] = "N/A"
                    
                    #Estado
                    try:
                        estado = fila.find_element(By.CSS_SELECTOR, "td:nth-child(6)").text.strip()
                        fila_datos["Estado"] = estado
                    except:
                        fila_datos["Estado"] = "N/A"
                    
                    datos_pagina.append(fila_datos)
                    
                except Exception as e:
                    logging.warning(f"Error procesando fila: {str(e)}")
                    continue
            
            return datos_pagina
            
        except Exception as e:
            logging.error(f"Error extrayendo datos de tabla: {str(e)}")
        return []
 
    def navegar_paginas(self):
        """
        Navega por todas las páginas de la tabla
        """
        try:
            pagina_actual = 1
            
            while True:
                logging.info(f"Procesando página {pagina_actual}")
                
                # Extraer datos de la página actual
                datos_pagina = self.extraer_datos_tabla_especifica()
                self.datos_extraidos.extend(datos_pagina)
                
                # Buscar botón "Siguiente"
                try:
                    boton_siguiente = self.driver.find_element(By.ID, "bus-table_next")
                    
                    # Verificar si el botón está habilitado
                    if "disabled" in boton_siguiente.get_attribute("class"):
                        logging.info("Llegamos a la última página")
                        break
                    
                    # Hacer click en siguiente
                    self.driver.execute_script("arguments[0].click();", boton_siguiente)
                    time.sleep(2)  # Esperar a que cargue la nueva página
                    
                    pagina_actual += 1
                    
                except Exception as e:
                    logging.info("No se encontró botón siguiente o error navegando")
                    break
            
            logging.info(f"Total de registros extraídos: {len(self.datos_extraidos)}")
            return True
            
        except Exception as e:
            logging.error(f"Error navegando páginas: {str(e)}")
            return False
    
    def crear_archivo_excel(self, nombre_archivo=None):
        """
        Crea un archivo Excel con los datos extraídos usando pandas y openpyxl,
        resaltando las empresas del sector agrícola y pecuario
        """
        try:
            if not self.datos_extraidos:
                logging.error("No hay datos para exportar")
                return False
            
            # Crear nombre del archivo si no se proporciona
            if not nombre_archivo:
                timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                nombre_archivo = f"datos_empresas_{timestamp}.xlsx"
            
            # Crear DataFrame de pandas
            df = pd.DataFrame(self.datos_extraidos)
            
            # Identificar empresas del sector agrícola y pecuario
            df['es_agropecuario'] = df['Actividad Económica'].apply(self._es_actividad_agropecuaria)
            
            # Crear archivo Excel con pandas
            with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Datos_Empresas', index=False)
            
            # Aplicar formato con openpyxl incluyendo resaltado
            self.aplicar_formato_excel_con_resaltado(nombre_archivo)
            
            # Contar empresas agropecuarias
            empresas_agropecuarias = df['es_agropecuario'].sum()
            total_empresas = len(df)
            
            logging.info(f"Archivo Excel creado exitosamente: {nombre_archivo}")
            print(f"✅ Archivo Excel creado: {nombre_archivo}")
            print(f"📊 Total de registros: {total_empresas}")
            print(f"🌾 Empresas agropecuarias: {empresas_agropecuarias} ({empresas_agropecuarias/total_empresas*100:.1f}%)")
            
            return True
            
        except Exception as e:
            logging.error(f"Error creando archivo Excel: {str(e)}")
            return False
    
    def _es_actividad_agropecuaria(self, actividad_economica):
        """
        Determina si una actividad económica pertenece al sector agrícola y pecuario
        basándose en códigos CIIU y palabras clave
        """
        if not actividad_economica or pd.isna(actividad_economica):
            return False
            
        actividad_texto = str(actividad_economica).upper().strip()
        
        # Lista de códigos CIIU del sector agrícola y pecuario (Sección A)
        codigos_agropecuarios = [
            # División 01 - Agricultura, ganadería, caza y actividades de servicios conexas
            '011', '0111', '0112', '0113', '0114', '0115', '0116', '0119',  # Cultivos no perennes
            '012', '0121', '0122', '0123', '0124', '0125', '0126', '0127', '0128', '0129',  # Cultivos perennes
            '013', '0130',  # Propagación de plantas
            '014', '0141', '0142', '0149',  # Ganadería
            '015', '0150',  # Explotación mixta
            '016', '0161', '0162', '0163', '0164',  # Actividades de apoyo a la agricultura y ganadería
            '017', '0170',  # Caza y actividades de servicios conexas
            
            # División 02 - Silvicultura y extracción de madera
            '021', '0210',  # Silvicultura y otras actividades forestales
            '022', '0220',  # Extracción de madera
            '023', '0230',  # Recolección de productos forestales diferentes a la madera
            '024', '0240',  # Servicios de apoyo a la silvicultura
            
            # División 03 - Pesca y acuicultura
            '031', '0311', '0312',  # Pesca
            '032', '0321', '0322',  # Acuicultura
        ]
        
        # Verificar códigos CIIU específicos
        for codigo in codigos_agropecuarios:
            if codigo in actividad_texto:
                return True
        
        # Palabras clave para actividades agropecuarias
        palabras_clave_agropecuarias = [
            # Agricultura y cultivos
            'AGRICULTURA', 'CULTIVO', 'SIEMBRA', 'COSECHA', 'PLANTACIÓN', 'INVERNADERO',
            'HORTICULTURA', 'FLORICULTURA', 'VIVERO', 'SEMILLAS', 'CEREALES', 'GRANOS',
            'ARROZ', 'MAÍZ', 'TRIGO', 'CEBADA', 'AVENA', 'SORGO', 'QUINUA',
            'LEGUMBRES', 'FRIJOL', 'LENTEJA', 'GARBANZO', 'SOYA', 'OLEAGINOSAS',
            'FRUTAS', 'FRUTALES', 'CÍTRICOS', 'BANANA', 'PLÁTANO', 'MANGO', 'AGUACATE',
            'VERDURAS', 'HORTALIZAS', 'TOMATE', 'PAPA', 'CEBOLLA', 'ZANAHORIA',
            'CAFÉ', 'CACAO', 'CAÑA DE AZÚCAR', 'ALGODÓN', 'TABACO',
            'FLORES', 'PLANTAS ORNAMENTALES', 'JARDINERÍA',
            
            # Ganadería
            'GANADERÍA', 'GANADERO', 'GANADO', 'BOVINO', 'VACUNO', 'LECHERÍA',
            'PORCINO', 'PORCICULTURA', 'CERDO', 'COCHINO', 'MARRANO',
            'AVICULTURA', 'AVÍCOLA', 'POLLO', 'GALLINA', 'HUEVOS', 'INCUBADORA',
            'OVINO', 'OVEJA', 'CORDERO', 'CAPRINO', 'CABRA',
            'EQUINO', 'CABALLO', 'YEGUA', 'MULAR', 'ASNAL',
            'PECUARIO', 'PECUARIA', 'CRÍA', 'CRIANZA', 'REPRODUCCIÓN ANIMAL',
            'VETERINARIA', 'ZOOTECNIA', 'ALIMENTACIÓN ANIMAL', 'FORRAJE',
            
            # Actividades mixtas y de apoyo
            'AGROPECUARIO', 'AGROPECUARIA', 'AGRÍCOLA', 'RURAL', 'CAMPO',
            'FINCA', 'HACIENDA', 'PREDIO', 'PARCELA', 'TERRENO AGRÍCOLA',
            'MAQUINARIA AGRÍCOLA', 'TRACTOR', 'COSECHADORA',
            'RIEGO', 'IRRIGACIÓN', 'FUMIGACIÓN', 'FERTILIZANTES', 'ABONOS',
            'INSECTICIDAS', 'HERBICIDAS', 'PLAGUICIDAS',
            'COOPERATIVA AGRÍCOLA', 'ASOCIACIÓN AGROPECUARIA',
            
            # Silvicultura
            'SILVICULTURA', 'FORESTAL', 'BOSQUE', 'MADERA', 'REFORESTACIÓN',
            'PLANTACIONES FORESTALES', 'TALA', 'EXTRACCIÓN MADERA',
            'PRODUCTOS FORESTALES', 'CAUCHO', 'RESINAS',
            
            # Pesca y acuicultura
            'PESCA', 'PESQUERO', 'PESCADO', 'ACUICULTURA', 'PISCICULTURA',
            'CAMARONERA', 'CAMARÓN', 'TRUCHA', 'TILAPIA', 'BAGRE',
            'MARINA', 'MARÍTIMA', 'FLUVIAL', 'ESTANQUE', 'JAULA'
        ]
        
        # Verificar palabras clave
        for palabra in palabras_clave_agropecuarias:
            if palabra in actividad_texto:
                return True
        
        return False
    
    def aplicar_formato_excel_con_resaltado(self, nombre_archivo):
        """
        Aplica formato profesional al archivo Excel, resaltando las empresas agropecuarias
        """
        try:
            # Cargar el archivo
            wb = load_workbook(nombre_archivo)
            ws = wb.active
            
            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Estilo para empresas agropecuarias
            agro_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
            agro_font = Font(bold=True, color="006400")  # Verde oscuro
            
            # Aplicar formato a encabezados
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            
            # Encontrar la columna de "es_agropecuario" (última columna)
            col_agropecuario = ws.max_column
            
            # Resaltar filas de empresas agropecuarias
            for row_num in range(2, ws.max_row + 1):  # Empezar desde fila 2 (después del encabezado)
                es_agropecuario_cell = ws.cell(row=row_num, column=col_agropecuario)
                
                if es_agropecuario_cell.value == True:
                    # Resaltar toda la fila
                    for col_num in range(1, ws.max_column):  # No incluir la columna auxiliar
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.fill = agro_fill
                        if col_num == 4:  # Columna "Actividad Económica" 
                            cell.font = agro_font
            
            # Ocultar la columna auxiliar "es_agropecuario"
            ws.column_dimensions[ws.cell(row=1, column=col_agropecuario).column_letter].hidden = True
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                if not column[0].column_letter == ws.cell(row=1, column=col_agropecuario).column_letter:  # No ajustar la columna oculta
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar bordes a toda la tabla (excepto columna oculta)
            from openpyxl.styles import Border, Side
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(max_col=ws.max_column-1):  # Excluir última columna
                for cell in row:
                    cell.border = thin_border
            
            # Agregar una leyenda
            leyenda_row = ws.max_row + 3
            ws.cell(row=leyenda_row, column=1, value="FILTRO POR SECTOR:")
            ws.cell(row=leyenda_row, column=1).font = Font(bold=True, size=12)
            
            ws.cell(row=leyenda_row + 1, column=1, value="Empresas del Sector Agrícola y Pecuario")
            ws.cell(row=leyenda_row + 1, column=1).fill = agro_fill
            ws.cell(row=leyenda_row + 1, column=1).font = agro_font
            
            # Guardar cambios
            wb.save(nombre_archivo)
            
            print(f"🎨 Formato aplicado: Empresas agropecuarias resaltadas en verde")
            
        except Exception as e:
            logging.error(f"Error aplicando formato: {str(e)}")
            print(f"⚠️ Error aplicando formato: {str(e)}")
            
    def navegar_a_empresas(self):
        """
        Navega específicamente a la página de empresas
        """
        try:
            print("Navegando a Empresas")
            # Variante 2: Buscar por clase y texto
            dropdown_empresa = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[@class='dropdown-toggle' and contains(text(), 'Empresa')]"))
            )  
            print("Encontrado")
            dropdown_empresa.click()
            
            time.sleep(2)
            # Esperar a que el submenú esté visible
            opcion_registro = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//ul[@class='dropdown-menu']//a[contains(text(), 'Registro y/o actualización de empresa')]"))
            )
            print("Opción encontrada")
            opcion_registro.click()
            
            time.sleep(2)
            
            # 5. Verificar que llegamos a la página correcta
            current_url = self.driver.current_url
            print(f"URL actual: {current_url}")
            
            if "funcionario/empresa" in current_url:
                print("✅ Navegación exitosa via menú")
                return True
            else:
                print("❌ No se pudo navegar via menú")
                return False
        except Exception as e:
            logging.error(f"Error navegando via menú: {str(e)}")
            print(f"❌ Error en navegación via menú: {str(e)}")
            return False
        
    def hacer_consulta_avanzada_con_fechas(self, fecha_desde, fecha_hasta):
        """
        Realiza consulta avanzada con rango de fechas para todo Cundinamarca
        """
        try:
            print("Haciendo búsqueda avanzada con rango de fechas")
            
            # Abrir modal de búsqueda avanzada
            boton_avanzada = self.wait.until(
                EC.element_to_be_clickable((By.ID, "btnBusqAvanzada"))
            )
            boton_avanzada.click()
            time.sleep(3)
            
            # Esperar a que el modal esté visible
            self.wait.until(
                EC.visibility_of_element_located((By.ID, "modalBusqAvanzada"))
            )
            
            # Seleccionar país (Colombia)
            input_pais = self.wait.until(EC.presence_of_element_located((By.ID, "bus-pais")))
            selector_pais = Select(input_pais)
            selector_pais.select_by_visible_text('Colombia')
            print("Se seleccionó Colombia")
            time.sleep(1)
            
            # Seleccionar departamento (Cundinamarca)
            input_departamento = self.wait.until(EC.presence_of_element_located((By.ID, "bus-departamento")))
            selector_departamento = Select(input_departamento)
            selector_departamento.select_by_visible_text('Cundinamarca')
            print("Se seleccionó Cundinamarca")
            time.sleep(1)
            
            # NO seleccionar municipio específico para obtener todo Cundinamarca
            
            # Configurar fecha DESDE con múltiples métodos
            self._configurar_fecha_datepicker("bus-fechaRDe-text", fecha_desde, "DESDE")
            
            # Configurar fecha HASTA con múltiples métodos  
            self._configurar_fecha_datepicker("bus-fechaRA-text", fecha_hasta, "HASTA")
            
            # Realizar búsqueda
            boton_buscar = self.wait.until(EC.element_to_be_clickable((By.ID, "btnBuscarAv")))
            boton_buscar.click()
            time.sleep(3)
            
            print("Búsqueda avanzada completada")
            return True
            
        except Exception as e:
            logging.error(f"Error en búsqueda avanzada con fechas: {str(e)}")
            print(f"Error en búsqueda avanzada: {str(e)}")
            return False
    
    def _configurar_fecha_datepicker(self, campo_id, fecha, tipo_fecha):
        """
        Configura fechas en campos con datepicker usando múltiples métodos
        """
        try:
            print(f"Configurando fecha {tipo_fecha}: {fecha}")
            
            # Método 1: Intentar escribir directamente en el campo
            try:
                input_fecha = self.wait.until(EC.presence_of_element_located((By.ID, campo_id)))
                
                # Limpiar el campo
                input_fecha.clear()
                time.sleep(0.5)
                
                # Usar JavaScript para establecer el valor directamente
                self.driver.execute_script(f"document.getElementById('{campo_id}').value = '{fecha}';")
                time.sleep(0.5)
                
                # Verificar si se estableció correctamente
                valor_actual = input_fecha.get_attribute('value')
                if valor_actual == fecha:
                    print(f"✅ Fecha {tipo_fecha} configurada correctamente: {fecha}")
                    return True
                else:
                    print(f"⚠️ Valor no coincide. Esperado: {fecha}, Actual: {valor_actual}")
                
            except Exception as e:
                print(f"❌ Método 1 falló para {tipo_fecha}: {str(e)}")
            
            # Método 2: Usar send_keys después de limpiar con JavaScript
            try:
                input_fecha = self.driver.find_element(By.ID, campo_id)
                
                # Limpiar con JavaScript
                self.driver.execute_script(f"document.getElementById('{campo_id}').value = '';")
                time.sleep(0.5)
                
                # Enfocar el campo
                input_fecha.click()
                time.sleep(0.5)
                
                # Enviar las teclas
                input_fecha.send_keys(fecha)
                time.sleep(0.5)
                
                # Verificar
                valor_actual = input_fecha.get_attribute('value')
                if valor_actual == fecha:
                    print(f"✅ Fecha {tipo_fecha} configurada con método 2: {fecha}")
                    return True
                    
            except Exception as e:
                print(f"❌ Método 2 falló para {tipo_fecha}: {str(e)}")
            
            # Método 3: Simular eventos de teclado
            try:
                from selenium.webdriver.common.keys import Keys
                
                input_fecha = self.driver.find_element(By.ID, campo_id)
                
                # Seleccionar todo y reemplazar
                input_fecha.click()
                time.sleep(0.5)
                input_fecha.send_keys(Keys.CONTROL + "a")
                time.sleep(0.5)
                input_fecha.send_keys(fecha)
                time.sleep(0.5)
                
                # Verificar
                valor_actual = input_fecha.get_attribute('value')
                if valor_actual == fecha:
                    print(f"✅ Fecha {tipo_fecha} configurada con método 3: {fecha}")
                    return True
                    
            except Exception as e:
                print(f"❌ Método 3 falló para {tipo_fecha}: {str(e)}")
            
            # Método 4: Disparar eventos JavaScript
            try:
                self.driver.execute_script(f"""
                    var input = document.getElementById('{campo_id}');
                    input.value = '{fecha}';
                    input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                    input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                """)
                time.sleep(1)
                
                input_fecha = self.driver.find_element(By.ID, campo_id)
                valor_actual = input_fecha.get_attribute('value')
                if valor_actual == fecha:
                    print(f"✅ Fecha {tipo_fecha} configurada con método 4: {fecha}")
                    return True
                    
            except Exception as e:
                print(f"❌ Método 4 falló para {tipo_fecha}: {str(e)}")
            
            print(f"❌ Todos los métodos fallaron para configurar fecha {tipo_fecha}")
            return False
            
        except Exception as e:
            logging.error(f"Error configurando fecha {tipo_fecha}: {str(e)}")
            print(f"❌ Error general configurando fecha {tipo_fecha}: {str(e)}")
            return False
    
    def ejecutar_proceso_completo_fechas(self, fecha_desde, fecha_hasta):
        """
        Ejecuta todo el proceso de extracción para un rango de fechas específico
        """
        try:
            print("🚀 Iniciando proceso de extracción de datos por rango de fechas...")
            
            if not self.configurar_driver():
                return False
            
            if not self.realizar_login():
                return False
            
            time.sleep(3)  # Dar tiempo para que cargue el dashboard
            
            # Navegar a la página de empresas
            if not self.navegar_a_empresas():
                return False
            
            print(f"\n--- Procesando datos del {fecha_desde} al {fecha_hasta} en Cundinamarca ---")
            
            # Limpiar los datos extraídos
            self.datos_extraidos = []
            
            if self.hacer_consulta_avanzada_con_fechas(fecha_desde, fecha_hasta):
                try:
                    # Esperar a que la tabla se actualice después del filtro
                    self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, TABLA_SELECTOR + " " + FILAS_DATOS_SELECTOR)))
                    print(f"📄 Tabla actualizada para el rango {fecha_desde} - {fecha_hasta}, extrayendo datos...")
                    
                    # Navegar por todas las páginas para extraer datos
                    self.navegar_paginas()
                    
                    # Crear el archivo Excel
                    if self.datos_extraidos:
                        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                        # Formato de fecha para nombre de archivo (reemplazar / con -)
                        fecha_desde_archivo = fecha_desde.replace('/', '-')
                        fecha_hasta_archivo = fecha_hasta.replace('/', '-')
                        nombre_archivo = f"empresas_cundinamarca_{fecha_desde_archivo}_a_{fecha_hasta_archivo}_{timestamp}.xlsx"
                        
                        if self.crear_archivo_excel(nombre_archivo):
                            print(f"✅ Proceso completado exitosamente!")
                            print(f"📁 Archivo generado: {nombre_archivo}")
                        else:
                            logging.warning("No se pudo crear el archivo Excel.")
                    else:
                        print(f"⚠️ No se extrajeron datos para el rango de fechas especificado")
                        logging.info(f"No se extrajeron datos para el rango {fecha_desde} - {fecha_hasta}")
                    
                except TimeoutException:
                    print(f"No se encontraron datos en la tabla para el rango de fechas especificado.")
                    logging.info(f"No se encontraron datos para el rango {fecha_desde} - {fecha_hasta}")
                except Exception as e:
                    print(f"Error al procesar la tabla: {e}")
                    logging.warning(f"Error al procesar la tabla: {e}")
            else:
                print("La búsqueda avanzada falló.")
            
            return True
            
        except Exception as e:
            logging.error(f"Error en proceso completo: {str(e)}")
            print(f"❌ Error en el proceso: {str(e)}")
            traceback.print_exc()
            return False
            
        finally:
            if self.driver:
                self.driver.quit()
                logging.info("Driver cerrado")

def main():
    """
    Función principal
    """
    try:
        # Configurar las fechas de búsqueda 
        # IMPORTANTE: Probar diferentes formatos según lo que acepte el sistema
        # Formatos comunes: "dd-MM-yyyy", "dd/MM/yyyy", "yyyy-MM-dd"
        fecha_desde = "01-08-2025"  # 1 de enero de 2024
        fecha_hasta = "01-09-2025"  # 1 de febrero de 2024
        
        print(f"🚀 Iniciando extracción de empresas de Cundinamarca")
        print(f"📅 Rango de fechas: {fecha_desde} - {fecha_hasta}")
        
        extractor = ExtractorDatosEmpresa()
        extractor.ejecutar_proceso_completo_fechas(fecha_desde, fecha_hasta)
        
    except KeyboardInterrupt:
        print("\n⚠️  Proceso interrumpido por el usuario")
        logging.info("Proceso interrumpido por el usuario")
    
    except Exception as e:
        print(f"❌ Error inesperado: {str(e)}")
        logging.error(f"Error inesperado: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    main()